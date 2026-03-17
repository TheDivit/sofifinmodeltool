from __future__ import annotations

# ============================================================
# SECTION 1: IMPORTS & CONFIGURATION
# ============================================================

import io
import json
import os
import re
import tempfile
import traceback
from typing import Any

import camelot
import pandas as pd
import requests
import streamlit as st
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PyPDF2 import PdfReader

load_dotenv()

st.set_page_config(
    page_title="FinStatement Parser",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================
# SECTION 2: CONSTANTS & PROVIDER REGISTRY
# ============================================================

LLM_PROVIDERS: dict[str, dict[str, Any]] = {
    "OpenAI": {
        "base_url": "https://api.openai.com/v1",
        "models_endpoint": "/models",
        "chat_endpoint": "/chat/completions",
        "auth_header": "Authorization",
        "auth_prefix": "Bearer ",
        "model_filter": lambda m: m.startswith(("gpt-4", "gpt-3.5", "o1", "o3", "o4")),
        "model_sort_priority": ["gpt-4o", "gpt-4-turbo", "o4-mini", "o3-mini"],
    },
    "Anthropic": {
        "base_url": "https://api.anthropic.com/v1",
        "models_endpoint": "/models",
        "chat_endpoint": "/messages",
        "auth_header": "x-api-key",
        "auth_prefix": "",
        "model_filter": lambda m: "claude" in m.lower(),
        "model_sort_priority": ["claude-sonnet-4", "claude-opus-4", "claude-haiku-4"],
    },
    "Google Gemini": {
        "base_url": "https://generativelanguage.googleapis.com/v1beta",
        "models_endpoint": "/models",
        "chat_endpoint": None,
        "auth_header": None,
        "auth_prefix": "",
        "model_filter": lambda m: "gemini" in m.lower(),
        "model_sort_priority": ["gemini-2.5-flash", "gemini-2.5-pro", "gemini-2.0-flash"],
    },
}

FALLBACK_MODELS: dict[str, list[str]] = {
    "OpenAI": ["gpt-4o", "gpt-4o-mini", "gpt-4-turbo", "o4-mini", "o3-mini"],
    "Anthropic": ["claude-sonnet-4-20250514", "claude-haiku-4-20250414"],
    "Google Gemini": ["gemini-2.5-flash", "gemini-2.5-pro", "gemini-2.0-flash"],
}

MAPPING_SYSTEM_PROMPT = """You are a financial data mapping specialist. Your job is to take raw extracted financial statement data (from PDF table extraction) and map it precisely into a provided template structure.

## YOUR TASK
You will receive:
1. One or more EXTRACTED TABLES (CSV format) from a financial statement PDF. These may contain Balance Sheet, Profit & Loss, and/or Cash Flow Statement data.
2. A TEMPLATE (CSV format) that defines the exact output structure required.

You must:
1. IDENTIFY what type of financial data each extracted table contains.
2. MAP each row/field from the extracted data to the correct row in the template.
3. TRANSFORM values as needed:
   - Standardise number formats (remove commas, handle parentheses as negatives, handle \"Cr\"/\"Dr\" notation).
   - Convert units if the extracted data uses different units (e.g., thousands vs lakhs vs crores).
   - Handle merged/split rows — sometimes one template row maps to multiple extracted rows (sum them), or one extracted row maps to multiple template rows.
4. PRESERVE the template's exact row labels and column structure.
5. Where data is not available in the extracted tables for a template row, use null.

## OUTPUT FORMAT
Return ONLY valid JSON (no markdown fencing, no explanation). The JSON must be:
{
  \"statement_type\": \"Balance Sheet | Profit & Loss | Cash Flow | Mixed\",
  \"data_unit\": \"Absolute | Thousands | Lakhs | Crores | Millions\",
  \"confidence_notes\": [\"list of any assumptions or low-confidence mappings\"],
  \"mapped_rows\": [
    {
      \"template_row_label\": \"exact label from template\",
      \"values\": {
        \"column_name_1\": value_or_null,
        \"column_name_2\": value_or_null
      },
      \"source_row\": \"original label from extracted data (for audit trail)\",
      \"confidence\": \"high | medium | low\"
    }
  ],
  \"unmapped_extracted_rows\": [\"list of extracted row labels that could not be mapped\"],
  \"unfilled_template_rows\": [\"list of template row labels with no matching data\"]
}

## CRITICAL RULES
- Numbers must be plain numbers (no commas, no currency symbols). Use negative numbers, not parentheses.
- If a value shows \"(1,234)\" or \"1,234 Cr\" or \"-1234\", normalise to -1234.
- If units differ between source and template, convert. State the conversion in confidence_notes.
- NEVER fabricate data. If unsure, set value to null and add a confidence_note.
- Return ONLY the JSON object. No other text."""


# ============================================================
# SECTION 3: SESSION STATE INITIALISATION
# ============================================================

DEFAULTS: dict[str, Any] = {
    "extracted_tables": [],
    "extracted_csvs": [],
    "template_csv": None,
    "template_df": None,
    "mapped_data": None,
    "llm_raw_response": None,
    "available_models": {},
    "processing_status": "",
    "parsing_reports": [],
    "camelot_settings": {
        "flavor": "lattice",
        "line_scale": 40,
        "split_text": True,
        "flag_size": True,
        "strip_text": "\n",
        "edge_tol": 50,
        "row_tol": 2,
    },
}

for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ============================================================
# SECTION 4: HELPER FUNCTIONS — LLM
# ============================================================

def _sort_models(models: list[str], priorities: list[str]) -> list[str]:
    """Sort model IDs with priority models first, then alphabetically.

    Args:
        models: Raw model ID list.
        priorities: Preferred model prefixes.

    Returns:
        Sorted model IDs.
    """
    unique_models = sorted(set(models))
    priority_bucket: list[str] = []
    non_priority_bucket: list[str] = []

    for model in unique_models:
        if any(model.startswith(p) for p in priorities):
            priority_bucket.append(model)
        else:
            non_priority_bucket.append(model)

    def priority_rank(model_name: str) -> int:
        for i, p in enumerate(priorities):
            if model_name.startswith(p):
                return i
        return len(priorities)

    priority_bucket.sort(key=lambda m: (priority_rank(m), m))
    non_priority_bucket.sort()
    return priority_bucket + non_priority_bucket


def fetch_available_models(provider: str, api_key: str) -> list[str]:
    """Fetch model IDs for the selected provider.

    Args:
        provider: Provider display name.
        api_key: Provider API key.

    Returns:
        List of filtered/sorted model IDs, or fallback models on failure.
    """
    if not api_key:
        st.warning("Please provide an API key before fetching models.")
        return FALLBACK_MODELS.get(provider, [])

    provider_cfg = LLM_PROVIDERS[provider]
    url = f"{provider_cfg['base_url']}{provider_cfg['models_endpoint']}"

    try:
        if provider == "Google Gemini":
            url = f"{url}?key={api_key}"
            response = requests.get(url, timeout=20)
        else:
            headers = {
                provider_cfg["auth_header"]: f"{provider_cfg['auth_prefix']}{api_key}",
            }
            if provider == "Anthropic":
                headers["anthropic-version"] = "2023-06-01"
            response = requests.get(url, headers=headers, timeout=20)

        response.raise_for_status()
        payload = response.json()

        model_ids: list[str] = []
        if provider in {"OpenAI", "Anthropic"}:
            for item in payload.get("data", []):
                model_id = item.get("id", "")
                if model_id:
                    model_ids.append(model_id)
        elif provider == "Google Gemini":
            for item in payload.get("models", []):
                name = item.get("name", "")
                model_id = name.split("/")[-1] if name else ""
                if model_id:
                    model_ids.append(model_id)

        filtered = [m for m in model_ids if provider_cfg["model_filter"](m)]
        if not filtered:
            raise ValueError("No matching models found in provider response.")

        sorted_models = _sort_models(filtered, provider_cfg["model_sort_priority"])
        st.session_state["available_models"][provider] = sorted_models
        return sorted_models
    except Exception as exc:  # pylint: disable=broad-except
        fallback = FALLBACK_MODELS.get(provider, [])
        st.warning(
            f"Model fetch failed for {provider}: {exc}. Using fallback model list instead."
        )
        st.session_state["available_models"][provider] = fallback
        return fallback


def call_llm(
    provider: str,
    model: str,
    api_key: str,
    system_prompt: str,
    user_prompt: str,
) -> str:
    """Call a selected LLM provider and return response text.

    Args:
        provider: Selected provider name.
        model: Selected model ID.
        api_key: Provider API key.
        system_prompt: System instructions.
        user_prompt: User content.

    Returns:
        Raw text output from the provider.
    """
    try:
        if provider == "OpenAI":
            from openai import OpenAI

            client = OpenAI(api_key=api_key)
            resp = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0,
                max_tokens=8192,
            )
            return resp.choices[0].message.content or ""

        if provider == "Anthropic":
            from anthropic import Anthropic

            client = Anthropic(api_key=api_key)
            resp = client.messages.create(
                model=model,
                system=system_prompt,
                messages=[{"role": "user", "content": user_prompt}],
                temperature=0,
                max_tokens=8192,
            )
            content_parts = []
            for block in resp.content:
                text = getattr(block, "text", None)
                if text:
                    content_parts.append(text)
            return "\n".join(content_parts).strip()

        if provider == "Google Gemini":
            import google.genai as genai

            client = genai.Client(api_key=api_key)
            merged_prompt = f"{system_prompt}\n\n{user_prompt}"
            resp = client.models.generate_content(
                model=model,
                contents=merged_prompt,
                config={"temperature": 0, "max_output_tokens": 8192},
            )
            return (getattr(resp, "text", "") or "").strip()

        raise ValueError(f"Unsupported provider: {provider}")
    except Exception as exc:  # pylint: disable=broad-except
        st.error(f"LLM call failed: {exc}")
        raise


def build_mapping_prompt(extracted_csvs: list[str], template_csv: str) -> str:
    """Build mapping prompt from extracted CSV tables and template CSV.

    Args:
        extracted_csvs: CSV strings from extracted PDF tables.
        template_csv: Target template CSV.

    Returns:
        Combined user prompt string.
    """
    parts = ["## EXTRACTED TABLES FROM PDF\n"]
    for i, csv_str in enumerate(extracted_csvs):
        parts.append(f"### Table {i + 1}\n```csv\n{csv_str}\n```\n")
    parts.append(f"## TEMPLATE (Target Structure)\n```csv\n{template_csv}\n```\n")
    parts.append("Map the extracted data into the template structure. Return ONLY JSON.")
    return "\n".join(parts)


def parse_llm_response(raw_response: str) -> dict[str, Any]:
    """Parse JSON response from LLM, handling common formatting wrappers.

    Args:
        raw_response: Raw LLM text.

    Returns:
        Parsed JSON object.
    """
    cleaned = raw_response.strip()

    fence_match = re.match(r"^```(?:json)?\s*(.*?)\s*```$", cleaned, re.DOTALL | re.IGNORECASE)
    if fence_match:
        cleaned = fence_match.group(1).strip()

    if cleaned.lower().startswith("json"):
        cleaned = cleaned[4:].strip()

    return json.loads(cleaned)


def validate_mapped_data_structure(mapped_data: dict[str, Any]) -> None:
    """Validate minimum schema expected from LLM output.

    Args:
        mapped_data: Parsed mapping payload.

    Raises:
        ValueError: If required keys are missing or malformed.
    """
    required_root = {
        "statement_type",
        "data_unit",
        "confidence_notes",
        "mapped_rows",
        "unmapped_extracted_rows",
        "unfilled_template_rows",
    }
    missing = required_root.difference(mapped_data.keys())
    if missing:
        raise ValueError(f"Missing required JSON keys: {sorted(missing)}")

    if not isinstance(mapped_data.get("mapped_rows"), list):
        raise ValueError("'mapped_rows' must be a list.")


# ============================================================
# SECTION 5: HELPER FUNCTIONS — CAMELOT EXTRACTION
# ============================================================

def extract_tables_from_pdf(
    pdf_path: str,
    pages: str,
    flavor: str = "lattice",
) -> tuple[list[pd.DataFrame], list[str]]:
    """Extract tables from specified pages of a PDF using Camelot.

    Args:
        pdf_path: Path to PDF file.
        pages: Page numbers/ranges like "1,3,5-7".
        flavor: Camelot flavor ("lattice" or "stream").

    Returns:
        A tuple containing:
            - list of pandas DataFrames
            - list of CSV strings
    """
    camelot_settings = st.session_state.get("camelot_settings", {})
    kwargs: dict[str, Any] = {
        "flavor": flavor,
        "split_text": bool(camelot_settings.get("split_text", True)),
        "flag_size": bool(camelot_settings.get("flag_size", True)),
        "strip_text": str(camelot_settings.get("strip_text", "\n")),
    }

    if flavor == "lattice":
        kwargs["line_scale"] = int(camelot_settings.get("line_scale", 40))
    else:
        kwargs["edge_tol"] = int(camelot_settings.get("edge_tol", 50))
        kwargs["row_tol"] = int(camelot_settings.get("row_tol", 2))

    tables = camelot.read_pdf(pdf_path, pages=pages, **kwargs)
    dataframes: list[pd.DataFrame] = []
    csvs: list[str] = []
    reports: list[dict[str, Any]] = []

    for table in tables:
        df = table.df
        csv_text = df.to_csv(index=False)
        dataframes.append(df)
        csvs.append(csv_text)
        reports.append(table.parsing_report)

    st.session_state["parsing_reports"] = reports
    return dataframes, csvs


# ============================================================
# SECTION 6: HELPER FUNCTIONS — EXCEL GENERATION
# ============================================================

def _normalize_numeric(value: Any) -> float | int | None:
    """Normalize mixed numeric string formats into numeric values.

    Args:
        value: Any value from mapped JSON.

    Returns:
        int/float if parseable, else None.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value

    text = str(value).strip()
    if text == "":
        return None

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1].strip()

    text = text.replace(",", "")
    text = re.sub(r"\s*(Cr|Dr)$", "", text, flags=re.IGNORECASE)
    text = text.replace("$", "").replace("₹", "")

    try:
        number = float(text)
        if negative:
            number = -number
        if number.is_integer():
            return int(number)
        return number
    except ValueError:
        return None


def generate_excel(mapped_data: dict[str, Any], template_df: pd.DataFrame) -> bytes:
    """Generate formatted Excel output from mapped data and template DataFrame.

    Args:
        mapped_data: LLM mapping JSON.
        template_df: Template DataFrame preserving output structure.

    Returns:
        Excel file content as bytes.
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to get active worksheet.")
    ws.title = "Financial Statement"

    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    total_top_border = Border(top=Side(style="thin", color="000000"))

    headers = list(template_df.columns)
    if len(headers) < 1:
        raise ValueError("Template must contain at least one column.")

    output_headers = headers + ["Notes"]

    for col_idx, header in enumerate(output_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="003366")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    label_col = headers[0]
    data_cols = headers[1:]

    mapped_lookup: dict[str, dict[str, Any]] = {}
    for row in mapped_data.get("mapped_rows", []):
        key = str(row.get("template_row_label", "")).strip().lower()
        if key:
            mapped_lookup[key] = row

    for row_offset, (_, template_row) in enumerate(template_df.iterrows(), start=2):
        excel_row = row_offset
        row_label = str(template_row[label_col])
        lookup_key = row_label.strip().lower()
        mapped_row = mapped_lookup.get(lookup_key, {})

        label_cell = ws.cell(row=excel_row, column=1, value=row_label)
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.border = thin_border

        values = mapped_row.get("values", {}) if isinstance(mapped_row, dict) else {}
        for j, col_name in enumerate(data_cols, start=2):
            raw_value = values.get(col_name)
            number = _normalize_numeric(raw_value)
            if number is None:
                cell = ws.cell(row=excel_row, column=j, value=None)
            else:
                cell = ws.cell(row=excel_row, column=j, value=number)
                cell.number_format = "#,##0.00" if isinstance(number, float) else "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if float(number) < 0:
                    cell.font = Font(color="FF0000")
            cell.border = thin_border

        confidence = mapped_row.get("confidence", "") if isinstance(mapped_row, dict) else ""
        notes_cell = ws.cell(row=excel_row, column=len(output_headers), value=confidence)
        notes_cell.alignment = Alignment(horizontal="center", vertical="center")
        notes_cell.border = thin_border

        label_lower = row_label.lower()
        if "total" in label_lower or "sub-total" in label_lower or "subtotal" in label_lower:
            for col_idx in range(1, len(output_headers) + 1):
                c = ws.cell(row=excel_row, column=col_idx)
                c.font = Font(bold=True, color=c.font.color.rgb if c.font and c.font.color else None)
                c.border = Border(
                    left=c.border.left,
                    right=c.border.right,
                    top=total_top_border.top,
                    bottom=c.border.bottom,
                )

    for col_idx, header in enumerate(output_headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            display = "" if value is None else str(value)
            max_len = max(max_len, len(display))
        ws.column_dimensions[chr(64 + col_idx)].width = min(max(12, int(max_len * 1.2)), 60)

    ws.freeze_panes = "B2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================
# SECTION 7: LLM PROMPT TEMPLATES
# ============================================================

@st.cache_data(show_spinner=False)
def parse_template_csv(template_bytes: bytes) -> tuple[pd.DataFrame, str]:
    """Parse uploaded template CSV bytes.

    Args:
        template_bytes: Raw CSV bytes from upload.

    Returns:
        Tuple of parsed DataFrame and UTF-8 decoded CSV string.
    """
    text = template_bytes.decode("utf-8", errors="replace")
    return pd.read_csv(io.StringIO(text)), text


def build_mapped_preview_df(mapped_data: dict[str, Any], template_df: pd.DataFrame) -> pd.DataFrame:
    """Build a table preview that follows template structure plus confidence.

    Args:
        mapped_data: Parsed mapping JSON.
        template_df: Template DataFrame.

    Returns:
        Preview DataFrame.
    """
    headers = list(template_df.columns)
    label_col = headers[0]
    data_cols = headers[1:]

    mapped_lookup: dict[str, dict[str, Any]] = {}
    for row in mapped_data.get("mapped_rows", []):
        key = str(row.get("template_row_label", "")).strip().lower()
        if key:
            mapped_lookup[key] = row

    rows_out: list[dict[str, Any]] = []
    for _, row in template_df.iterrows():
        label = str(row[label_col])
        mapped = mapped_lookup.get(label.strip().lower(), {})
        values = mapped.get("values", {}) if isinstance(mapped, dict) else {}

        out: dict[str, Any] = {label_col: label}
        for c in data_cols:
            out[c] = values.get(c)
        out["Notes"] = mapped.get("confidence") if isinstance(mapped, dict) else None
        rows_out.append(out)

    return pd.DataFrame(rows_out)


def summarize_confidence(mapped_data: dict[str, Any]) -> dict[str, int]:
    """Count confidence classes from mapped rows.

    Args:
        mapped_data: Parsed mapping JSON.

    Returns:
        Dict with high/medium/low counts.
    """
    counts = {"high": 0, "medium": 0, "low": 0}
    for row in mapped_data.get("mapped_rows", []):
        conf = str(row.get("confidence", "")).strip().lower()
        if conf in counts:
            counts[conf] += 1
    return counts


# ============================================================
# SECTION 8: STREAMLIT UI — SIDEBAR
# ============================================================

st.title("FinStatement Parser")
st.caption("Extract financial statement tables from PDF, map to template CSV with an LLM, and export Excel.")

with st.sidebar:
    st.header("🤖 LLM Configuration")

    provider = st.selectbox("Provider", options=["OpenAI", "Anthropic", "Google Gemini"])

    env_key_name = {
        "OpenAI": "OPENAI_API_KEY",
        "Anthropic": "ANTHROPIC_API_KEY",
        "Google Gemini": "GOOGLE_API_KEY",
    }[provider]
    default_key = os.getenv(env_key_name, "")

    api_key = st.text_input(
        "API Key",
        type="password",
        value=default_key,
        help="Your key is not stored anywhere.",
    )

    current_models = st.session_state["available_models"].get(provider, FALLBACK_MODELS[provider])

    fetch_col, refresh_col = st.columns([2, 1])
    with fetch_col:
        if st.button("Fetch Models", use_container_width=True):
            with st.spinner("Fetching models..."):
                current_models = fetch_available_models(provider, api_key)
    with refresh_col:
        if st.button("Reset", use_container_width=True):
            st.session_state["available_models"][provider] = FALLBACK_MODELS[provider]
            current_models = FALLBACK_MODELS[provider]

    model = st.selectbox(
        "Model",
        options=current_models,
        index=0 if current_models else None,
        disabled=not current_models,
    )

    with st.expander("⚙️ Camelot Settings", expanded=False):
        flavor = st.selectbox("Flavor", options=["lattice", "stream"], index=0)
        line_scale = st.slider("Line Scale", min_value=15, max_value=150, value=40)
        split_text = st.checkbox("Split Text", value=True)
        flag_size = st.checkbox("Flag Size", value=True)
        strip_text = st.text_input("Strip Text", value="\\n")
        edge_tol = st.slider("Edge Tolerance", min_value=0, max_value=500, value=50)
        row_tol = st.slider("Row Tolerance", min_value=0, max_value=50, value=2)

        st.session_state["camelot_settings"] = {
            "flavor": flavor,
            "line_scale": line_scale,
            "split_text": split_text,
            "flag_size": flag_size,
            "strip_text": strip_text,
            "edge_tol": edge_tol,
            "row_tol": row_tol,
        }

        if flavor == "lattice":
            st.info("Lattice works best for bordered tables.")
        else:
            st.info("Stream works best for borderless tables and text-aligned tables.")


# ============================================================
# SECTION 9: STREAMLIT UI — MAIN CONTENT
# ============================================================

step1, step2, step3 = st.tabs(
    ["📄 Step 1: Upload Files", "📋 Step 2: Review Extracted Data", "✅ Step 3: Mapped Output"]
)

with step1:
    upload_col1, upload_col2 = st.columns(2)
    with upload_col1:
        pdf_upload = st.file_uploader("Upload PDF", type=["pdf"], key="pdf_uploader")
    with upload_col2:
        template_upload = st.file_uploader("Upload Template CSV", type=["csv"], key="template_uploader")

    page_numbers = st.text_input("Page numbers", placeholder="e.g. 1,3,5-7")

    if template_upload is not None:
        try:
            template_df, template_csv = parse_template_csv(template_upload.getvalue())
            st.session_state["template_df"] = template_df
            st.session_state["template_csv"] = template_csv
            st.success("Template CSV loaded successfully.")
        except Exception as exc:  # pylint: disable=broad-except
            st.error(f"Failed to parse template CSV: {exc}")
            with st.expander("🔍 Debug: Template Parse Traceback"):
                st.code(traceback.format_exc(), language="text")

    if st.button("Extract Tables ▶", type="primary", use_container_width=False):
        if pdf_upload is None:
            st.error("Please upload a PDF file.")
        elif not page_numbers.strip():
            st.error("Please enter page numbers (example: 1,3,5-7).")
        elif st.session_state["template_df"] is None:
            st.error("Please upload a template CSV before extraction.")
        else:
            tmp_path = ""
            try:
                with st.spinner("Extracting tables with Camelot..."):
                    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                        tmp.write(pdf_upload.getvalue())
                        tmp_path = tmp.name

                    try:
                        reader = PdfReader(tmp_path)
                        page_count = len(reader.pages)
                        st.info(f"PDF loaded with {page_count} pages. Camelot pages are 1-indexed.")
                    except Exception:
                        pass

                    flavor_value = st.session_state["camelot_settings"].get("flavor", "lattice")
                    tables, csvs = extract_tables_from_pdf(tmp_path, page_numbers.strip(), flavor=flavor_value)
                    st.session_state["extracted_tables"] = tables
                    st.session_state["extracted_csvs"] = csvs
                    st.session_state["mapped_data"] = None

                if not tables:
                    st.warning("No tables were extracted. Verify page numbers, then try 'stream' flavor if needed.")
                    st.info("Camelot works on text-based PDFs. If this is a scanned PDF, run OCR first.")
                else:
                    st.success(f"Extracted {len(tables)} table(s).")
            except Exception as exc:  # pylint: disable=broad-except
                st.error(f"Failed to extract tables: {exc}")
                with st.expander("🔍 Debug: Extraction Traceback"):
                    st.code(traceback.format_exc(), language="text")
                if "ghostscript" in str(exc).lower():
                    st.error("Ghostscript appears missing. Install it at OS level and retry.")
            finally:
                if tmp_path and os.path.exists(tmp_path):
                    try:
                        os.unlink(tmp_path)
                    except OSError:
                        pass

with step2:
    extracted_tables = st.session_state["extracted_tables"]
    reports = st.session_state.get("parsing_reports", [])

    if extracted_tables:
        for i, table_df in enumerate(extracted_tables, start=1):
            report = reports[i - 1] if i - 1 < len(reports) else {}
            accuracy = report.get("accuracy", "n/a")
            page = report.get("page", "?")
            whitespace = report.get("whitespace", "n/a")

            st.markdown(f"**Table {i} (Page {page})**")
            st.caption(f"Accuracy: {accuracy}% | Whitespace: {whitespace}")
            st.dataframe(table_df, use_container_width=True, height=250)

            if isinstance(accuracy, (float, int)) and float(accuracy) < 85:
                st.warning(
                    "Low extraction accuracy detected. Consider changing Camelot flavor/settings and retrying."
                )

        if st.session_state["template_df"] is not None:
            st.markdown("**Template Preview**")
            st.dataframe(st.session_state["template_df"], use_container_width=True, height=250)

        if st.button("Map to Template ▶", type="primary"):
            if not api_key:
                st.error("Please provide an API key in the sidebar.")
            elif not model:
                st.error("Please select a model.")
            elif st.session_state["template_csv"] is None:
                st.error("Template CSV is missing. Re-upload the template in Step 1.")
            elif not st.session_state["extracted_csvs"]:
                st.error("No extracted table CSVs found. Extract tables first.")
            else:
                user_prompt = build_mapping_prompt(
                    st.session_state["extracted_csvs"],
                    st.session_state["template_csv"],
                )

                last_error = None
                with st.spinner("Mapping extracted data to template with LLM..."):
                    for attempt in range(1, 4):
                        try:
                            raw_response = call_llm(
                                provider=provider,
                                model=model,
                                api_key=api_key,
                                system_prompt=MAPPING_SYSTEM_PROMPT,
                                user_prompt=user_prompt,
                            )
                            st.session_state["llm_raw_response"] = raw_response
                            mapped = parse_llm_response(raw_response)
                            validate_mapped_data_structure(mapped)
                            st.session_state["mapped_data"] = mapped
                            st.success(f"Mapping completed successfully on attempt {attempt}.")
                            last_error = None
                            break
                        except Exception as exc:  # pylint: disable=broad-except
                            last_error = exc

                if last_error is not None:
                    st.error(f"Failed to parse/validate LLM output after retries: {last_error}")
                    st.warning("Try a different model or re-run mapping. Review raw response in debug section.")
    else:
        st.info("Extract tables in Step 1 to review them here.")

with step3:
    mapped_data = st.session_state["mapped_data"]
    template_df = st.session_state["template_df"]

    if mapped_data and template_df is not None:
        counts = summarize_confidence(mapped_data)
        metric_c1, metric_c2, metric_c3 = st.columns(3)
        metric_c1.metric("🟢 High", counts["high"])
        metric_c2.metric("🟡 Medium", counts["medium"])
        metric_c3.metric("🔴 Low", counts["low"])

        confidence_notes = mapped_data.get("confidence_notes", [])
        if confidence_notes:
            st.markdown("**Confidence Notes**")
            for note in confidence_notes:
                st.write(f"- {note}")

        unmapped = mapped_data.get("unmapped_extracted_rows", [])
        unfilled = mapped_data.get("unfilled_template_rows", [])
        if unmapped:
            st.warning(f"Unmapped extracted rows: {', '.join(map(str, unmapped))}")
        if unfilled:
            st.warning(f"Template rows with no data: {', '.join(map(str, unfilled))}")

        preview_df = build_mapped_preview_df(mapped_data, template_df)
        st.markdown("**Mapped Data Preview**")
        st.dataframe(preview_df, use_container_width=True, height=350)

        try:
            excel_bytes = generate_excel(mapped_data, template_df)
            d1, d2 = st.columns([2, 1])
            with d1:
                st.download_button(
                    label="⬇ Download Excel",
                    data=excel_bytes,
                    file_name="finstatement_mapped_output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with d2:
                if st.button("🔄 Re-run Mapping", use_container_width=True):
                    st.session_state["mapped_data"] = None
                    st.rerun()
        except Exception as exc:  # pylint: disable=broad-except
            st.error(f"Excel generation failed: {exc}")
            with st.expander("🔍 Debug: Excel Traceback"):
                st.code(traceback.format_exc(), language="text")
    else:
        st.info("Run mapping in Step 2 to view mapped output and download Excel.")

with st.expander("🔍 Debug: Raw LLM Response", expanded=False):
    raw_resp = st.session_state.get("llm_raw_response")
    if raw_resp:
        st.code(raw_resp, language="json")
    else:
        st.write("No LLM response yet.")
