#!/usr/bin/env python3
"""
Comprehensive smoke + integration test for sofifinmodeltool.

Tests:
  1. Ghostscript detection patch
  2. PDF creation (fpdf2)
  3. Camelot table extraction (lattice, then stream fallback)
  4. Numeric parsing (commas, parens-negative, dash, nil, currencies)
  5. pick_best_table heuristic
  6. detect_numeric_columns
  7. detect_scale_from_header
  8. Excel template creation, fuzzy sheet matching, and write_to_template
  9. pdfplumber keyword page scanning
"""
import sys
import os

# ---------------------------------------------------------------------------
# 0. Ghostscript detection patch (same as app.py startup)
# ---------------------------------------------------------------------------
if sys.platform == "darwin":
    import ctypes.util
    if ctypes.util.find_library("gs") is None:
        gs_lib = "/opt/homebrew/lib/libgs.dylib"
        if os.path.exists(gs_lib):
            os.environ.setdefault("DYLD_FALLBACK_LIBRARY_PATH", "/opt/homebrew/lib")
            import camelot.backends.ghostscript_backend as _gsb
            _gsb.installed_posix = lambda: True
            print("[fix] Patched ghostscript detection for Homebrew on macOS ARM")

# Make sure app is importable
sys.path.insert(0, os.path.dirname(__file__))

passed = 0
failed = 0

def check(label, condition):
    global passed, failed
    if condition:
        print(f"  ✅ {label}")
        passed += 1
    else:
        print(f"  ❌ {label}")
        failed += 1

# ---------------------------------------------------------------------------
# 1. Create test PDF
# ---------------------------------------------------------------------------
print("\n=== 1. Creating test PDF ===")
try:
    from fpdf import FPDF
except ImportError:
    os.system(f"{sys.executable} -m pip install fpdf2 -q")
    from fpdf import FPDF

PDF_PATH = "/tmp/sofi_test_table.pdf"

pdf = FPDF()
pdf.add_page()
pdf.set_font("Helvetica", "B", 14)
pdf.cell(0, 10, "Income Statement", new_x="LMARGIN", new_y="NEXT", align="C")
pdf.ln(3)
pdf.set_font("Helvetica", "I", 9)
pdf.cell(0, 6, "(Rs. in Crores)", new_x="LMARGIN", new_y="NEXT", align="C")
pdf.ln(3)
pdf.set_font("Helvetica", "", 10)
data = [
    ["Item", "FY2023", "FY2022"],
    ["Revenue", "10,000", "8,500"],
    ["Cost of Goods Sold", "(6,000)", "(5,100)"],
    ["Gross Profit", "4,000", "3,400"],
    ["Operating Expenses", "(2,000)", "(1,800)"],
    ["Net Profit", "2,000", "1,600"],
]
col_widths = [60, 35, 35]
for row in data:
    for j, cell in enumerate(row):
        pdf.cell(col_widths[j], 8, cell, border=1)
    pdf.ln()
pdf.output(PDF_PATH)
check("PDF created", os.path.exists(PDF_PATH))

# ---------------------------------------------------------------------------
# 2. Camelot extraction
# ---------------------------------------------------------------------------
print("\n=== 2. Camelot extraction (lattice) ===")
import camelot
tables = camelot.read_pdf(PDF_PATH, pages="1", flavor="lattice")
check(f"Lattice found {len(tables)} table(s)", len(tables) >= 1)

if len(tables) == 0:
    print("  Falling back to stream…")
    tables = camelot.read_pdf(PDF_PATH, pages="1", flavor="stream")
    check(f"Stream found {len(tables)} table(s)", len(tables) >= 1)

df = tables[0].df
print(f"  Shape: {df.shape}")
print(df.to_string(index=False))
check("Table has 6 rows", df.shape[0] == 6)
check("Table has 3 columns", df.shape[1] == 3)

# ---------------------------------------------------------------------------
# 3. Numeric parsing
# ---------------------------------------------------------------------------
print("\n=== 3. Numeric parsing ===")
from app import parse_numeric

cases = [
    ("10,000", 10000.0),
    ("(6,000)", -6000.0),
    ("-", 0.0),
    ("3,400", 3400.0),
    ("nil", 0.0),
    ("₹1,234.56", 1234.56),
    ("$100", 100.0),
    ("Revenue", None),   # non-numeric label
]
for raw, expected in cases:
    result = parse_numeric(raw)
    check(f"parse_numeric({raw!r}) == {expected}", result == expected)

# ---------------------------------------------------------------------------
# 4. detect_numeric_columns
# ---------------------------------------------------------------------------
print("\n=== 4. detect_numeric_columns ===")
from app import detect_numeric_columns
num_cols = detect_numeric_columns(df)
check(f"Numeric columns {num_cols} include col 1 and 2", 1 in num_cols and 2 in num_cols)
check("Column 0 (labels) NOT numeric", 0 not in num_cols)

# ---------------------------------------------------------------------------
# 5. pick_best_table
# ---------------------------------------------------------------------------
print("\n=== 5. pick_best_table ===")
import pandas as pd
from app import pick_best_table

small = pd.DataFrame({"a": ["x"], "b": ["1"]})
big = df.copy()
best = pick_best_table([small, big])
check("Picks bigger table", best is not None and len(best) == len(big))

# Concat case — same column count
frag1 = df.iloc[:3].copy()
frag2 = df.iloc[3:].copy()
merged = pick_best_table([frag1, frag2])
check("Merges same-structure fragments", merged is not None and len(merged) == len(df))

# ---------------------------------------------------------------------------
# 6. detect_scale_from_header
# ---------------------------------------------------------------------------
print("\n=== 6. detect_scale_from_header ===")
from app import detect_scale_from_header

# Build a dataframe that mimics having "(Rs. in Crores)" in the header area
header_df = pd.DataFrame([
    ["Income Statement", "", ""],
    ["(Rs. in Crores)", "", ""],
    ["Item", "FY2023", "FY2022"],
    ["Revenue", "10,000", "8,500"],
])
scale = detect_scale_from_header(header_df)
check(f"Scale detected as '{scale}'", scale == "Crores")

# ---------------------------------------------------------------------------
# 7. Excel template creation + fuzzy sheet match + write_to_template
# ---------------------------------------------------------------------------
print("\n=== 7. Excel template + write_to_template ===")
from openpyxl import Workbook, load_workbook
import io
from app import fuzzy_match_sheet, write_to_template

# Create a mock Excel template in memory
wb = Workbook()
ws = wb.active
ws.title = "Income Stmt"  # fuzzy match for "P&L / Income Statement"
labels = ["Revenue", "COGS", "Gross Profit", "OpEx", "Net Profit", "EBITDA"]
for i, lbl in enumerate(labels, start=1):
    ws.cell(row=i, column=1, value=lbl)
buf = io.BytesIO()
wb.save(buf)
tpl_bytes = buf.getvalue()

# Test fuzzy match
sheet_names = ["Income Stmt"]
match = fuzzy_match_sheet(sheet_names, "P&L / Income Statement")
check(f"Fuzzy matched '{match}'", match == "Income Stmt")

# Test write_to_template
results = {
    "P&L / Income Statement": {
        "statement_type": "P&L / Income Statement",
        "page": 1,
        "mapping": {
            "Revenue": "Revenue",
            "COGS": "Cost of Goods Sold",
            "Gross Profit": "Gross Profit",
            "OpEx": "Operating Expenses",
            "Net Profit": "Net Profit",
            "EBITDA": None,   # no match — should stay blank
        },
        "values": {
            "Revenue": [10000.0, 8500.0],
            "Cost of Goods Sold": [-6000.0, -5100.0],
            "Gross Profit": [4000.0, 3400.0],
            "Operating Expenses": [-2000.0, -1800.0],
            "Net Profit": [2000.0, 1600.0],
        },
        "scale": "Crores",
    }
}
out_bytes = write_to_template(tpl_bytes, results, "test_report.pdf")
check("write_to_template returned bytes", len(out_bytes) > 0)

# Verify the written workbook
out_wb = load_workbook(io.BytesIO(out_bytes))
out_ws = out_wb["Income Stmt"]
rev_val = out_ws.cell(row=1, column=2).value  # Revenue → FY2023
check(f"Revenue cell = {rev_val}", rev_val == 10000.0)
ebitda_val = out_ws.cell(row=6, column=2).value  # EBITDA → should be None/blank
check(f"EBITDA cell = {ebitda_val} (blank)", ebitda_val is None)

# Check metadata row exists
found_meta = False
for row in range(1, out_ws.max_row + 1):
    v = out_ws.cell(row=row, column=1).value
    if v and "test_report.pdf" in str(v):
        found_meta = True
        break
check("Source metadata written", found_meta)

# ---------------------------------------------------------------------------
# 8. pdfplumber keyword scanning
# ---------------------------------------------------------------------------
print("\n=== 8. pdfplumber keyword scanning ===")
from app import find_pages_for_statement

pages = find_pages_for_statement(PDF_PATH, "P&L / Income Statement")
check(f"Keyword scan found pages: {pages}", 1 in pages)

pages_bs = find_pages_for_statement(PDF_PATH, "Balance Sheet")
check(f"Balance Sheet not on this PDF: {pages_bs}", len(pages_bs) == 0)

# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------
print(f"\n{'='*50}")
print(f"Results: {passed} passed, {failed} failed")
if failed:
    print("⚠️  Some tests failed!")
    sys.exit(1)
else:
    print("✅ All tests passed!")
    sys.exit(0)
